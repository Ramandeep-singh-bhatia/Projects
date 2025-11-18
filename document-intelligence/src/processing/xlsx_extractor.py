"""
XLSX (Excel) document extractor.
Handles Excel spreadsheets with table awareness.
"""

from pathlib import Path
from typing import Dict, Any
from openpyxl import load_workbook
from src.processing.base_extractor import BaseExtractor, ExtractedDocument
from src.utils.logger import app_logger as logger


class XLSXExtractor(BaseExtractor):
    """Extractor for XLSX documents."""

    def can_handle(self, file_path: Path) -> bool:
        """Check if file is an XLSX."""
        return self._get_file_extension(file_path) in ['xlsx', 'xls']

    def extract(self, file_path: Path) -> ExtractedDocument:
        """
        Extract content and metadata from XLSX.

        Args:
            file_path: Path to XLSX file

        Returns:
            ExtractedDocument with content and metadata
        """
        logger.info(f"Extracting XLSX: {file_path}")

        try:
            workbook = load_workbook(file_path, data_only=True)

            # Extract content from all sheets
            content = self._extract_content(workbook)

            # Extract metadata
            metadata = self._extract_metadata(workbook, file_path)

            word_count = self._count_words(content)

            logger.info(f"Successfully extracted XLSX: {len(workbook.sheetnames)} sheets")

            return ExtractedDocument(
                content=content,
                metadata=metadata,
                page_count=len(workbook.sheetnames),
                word_count=word_count,
                language="en",
                title=metadata.get('title'),
                author=metadata.get('author'),
                created_date=metadata.get('created_date'),
                modified_date=metadata.get('modified_date')
            )

        except Exception as e:
            logger.error(f"Failed to extract XLSX {file_path}: {str(e)}")
            raise

    def _extract_content(self, workbook) -> str:
        """Extract content from all sheets."""
        content = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content.append(f"\n=== Sheet: {sheet_name} ===\n")

            # Extract data rows
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if any(cell is not None for cell in row):
                    row_text = "\t".join(str(cell or '') for cell in row)
                    rows_data.append(row_text)

            content.append("\n".join(rows_data))

        return "\n".join(content)

    def _extract_metadata(self, workbook, file_path: Path) -> Dict[str, Any]:
        """Extract XLSX metadata."""
        metadata = {}

        try:
            # Workbook properties
            props = workbook.properties

            metadata['title'] = props.title or ""
            metadata['author'] = props.creator or ""
            metadata['subject'] = props.subject or ""
            metadata['keywords'] = props.keywords or ""
            metadata['category'] = props.category or ""
            metadata['comments'] = props.description or ""

            # Dates
            if props.created:
                metadata['created_date'] = props.created
            if props.modified:
                metadata['modified_date'] = props.modified

            # Sheet information
            metadata['sheet_count'] = len(workbook.sheetnames)
            metadata['sheet_names'] = workbook.sheetnames

        except Exception as e:
            logger.warning(f"Failed to extract XLSX metadata: {str(e)}")

        return metadata
